import openpyxl
import pytest

class Xluti:

    @staticmethod
    def getRow(File,sheetname):
        workBook=openpyxl.load_workbook(File)
        sheet=workBook[sheetname]
        return sheet.max_row

    @staticmethod
    def getColumn(File, sheetname):
        workBook = openpyxl.load_workbook(File)
        sheet = workBook[sheetname]
        return sheet.max_column

    @staticmethod
    def getRead(File, sheetname,row,column):
        workBook = openpyxl.load_workbook(File)
        sheet = workBook[sheetname]
        return sheet.cell(row,column).value

    @staticmethod
    def getWright(File, sheetname,row,column,data):
        workBook = openpyxl.load_workbook(File)
        sheet = workBook[sheetname]
        (sheet.cell(row,column).value)=data
        workBook.save(File)
